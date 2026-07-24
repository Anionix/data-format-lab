from __future__ import annotations

import csv
import hashlib
import io
import platform
import zipfile
from collections.abc import Iterable, Mapping
from datetime import date, datetime
from pathlib import Path

from .contracts import normalized_columns
from .json_contract import strict_json_dumps
from .nyc_snapshot import CaptureState, fail_active_capture, finalize_capture, nyc_rows


_GEONAMES_COLUMNS = (
    "geonameid", "name", "asciiname", "alternatenames", "latitude",
    "longitude", "feature_class", "feature_code", "country_code",
    "cc2", "admin1_code", "admin2_code", "admin3_code", "admin4_code",
    "population", "elevation", "dem", "timezone", "modification_date",
)


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat(sep=" ") if isinstance(value, datetime) else value.isoformat()
    return str(value)


def _columns(manifest: Mapping[str, object]) -> tuple[str, ...]:
    return tuple(column["name"] for column in normalized_columns(manifest.get("columns")))


def _types(manifest: Mapping[str, object]) -> dict[str, str]:
    return {
        column["name"]: column["arrow_type"]
        for column in normalized_columns(manifest.get("columns"))
    }


def _write_rows(destination: Path, manifest: Mapping[str, object], rows: Iterable[Mapping[str, object]]) -> int:
    names = _columns(manifest)
    types = _types(manifest)
    count = 0
    with destination.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=names, lineterminator="\n")
        writer.writeheader()
        for row in rows:
            output: dict[str, str] = {}
            for name in names:
                value = row.get(name)
                if value is None or value == "":
                    output[name] = ""
                elif types[name] == "bool":
                    output[name] = "true" if value is True or str(value).lower() == "true" else "false"
                else:
                    output[name] = _text(value)
            writer.writerow(output)
            count += 1
    return count


def _bank_rows(raw: bytes, names: tuple[str, ...]) -> Iterable[Mapping[str, object]]:
    with zipfile.ZipFile(io.BytesIO(raw)) as outer:
        member = next(name for name in outer.namelist() if name.endswith("bank-additional.zip"))
        nested = outer.read(member)
    with zipfile.ZipFile(io.BytesIO(nested)) as inner:
        member = next(name for name in inner.namelist() if name.endswith("bank-additional-full.csv"))
        data = io.TextIOWrapper(inner.open(member), encoding="latin-1", newline="")
        with data:
            for row in csv.DictReader(data, delimiter=";"):
                yield row


def _retail_rows(raw: bytes) -> Iterable[Mapping[str, object]]:
    from openpyxl import load_workbook

    with zipfile.ZipFile(io.BytesIO(raw)) as archive:
        member = next(name for name in archive.namelist() if name.endswith("online_retail_II.xlsx"))
        workbook_bytes = archive.read(member)
    workbook = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            iterator = sheet.iter_rows(values_only=True)
            source_names = ["Customer_ID" if value == "Customer ID" else str(value) for value in next(iterator)]
            for values in iterator:
                if not any(value is not None for value in values):
                    continue
                yield dict(zip(source_names, values, strict=False))
    finally:
        workbook.close()


def _geonames_rows(raw: bytes) -> Iterable[Mapping[str, object]]:
    records: list[tuple[int, tuple[str, ...]]] = []
    with zipfile.ZipFile(io.BytesIO(raw)) as archive:
        member = next(name for name in archive.namelist() if name.endswith("cities500.txt"))
        data = io.TextIOWrapper(archive.open(member), encoding="utf-8", newline="")
        with data:
            for line_number, line in enumerate(data, start=1):
                values = tuple(line.rstrip("\r\n").split("\t"))
                if len(values) != len(_GEONAMES_COLUMNS):
                    raise ValueError(f"GeoNames row {line_number} has {len(values)} fields, expected 19")
                try:
                    geonameid = int(values[0])
                except ValueError as error:
                    raise ValueError(f"GeoNames row {line_number} has invalid geonameid {values[0]!r}") from error
                records.append((geonameid, values))

    records.sort()
    for _, values in records:
        yield dict(zip(_GEONAMES_COLUMNS, values, strict=True))


def _owid_rows(raw: bytes) -> Iterable[Mapping[str, object]]:
    text = io.TextIOWrapper(io.BytesIO(raw), encoding="utf-8", newline="")
    with text:
        yield from csv.DictReader(text)


def materialize_official(
    dataset_id: str,
    manifest: Mapping[str, object],
    raw: bytes,
    destination: Path,
) -> Path:
    destination.mkdir(parents=True, exist_ok=False)
    if raw:
        (destination / "raw.bin").write_bytes(raw)
    source_format = str(manifest.get("source_format", ""))
    output = destination / "source.csv"
    if dataset_id == "uci-bank-marketing":
        rows = _bank_rows(raw, _columns(manifest))
    elif dataset_id == "uci-online-retail-ii":
        rows = _retail_rows(raw)
    elif dataset_id == "geonames-cities500":
        rows = _geonames_rows(raw)
    elif dataset_id == "owid-energy":
        rows = _owid_rows(raw)
    elif dataset_id == "nyc-311-2010-2019":
        rows = nyc_rows(manifest, destination)
    else:
        raise ValueError(f"no official normalizer for {dataset_id} ({source_format})")
    partial = destination / "source.csv.partial"
    is_nyc = dataset_id == "nyc-311-2010-2019"
    materialized = partial if is_nyc else output
    succeeded = False
    try:
        row_count = _write_rows(materialized, manifest, rows)
        if is_nyc:
            partial.replace(output)
        materialization = destination / "materialization.json"
        with output.open("rb") as source:
            source_sha256 = hashlib.file_digest(source, "sha256").hexdigest()
        materialization.write_text(
            strict_json_dumps(
                {
                    "dataset_id": dataset_id,
                    "rows": row_count,
                    "source_format": source_format,
                    "source_sha256": source_sha256,
                    "writer": {
                        "format": "csv",
                        "encoding": "utf-8",
                        "newline": "",
                        "dialect": "excel",
                        "lineterminator": "\n",
                    },
                    "runtime_versions": {"python": platform.python_version()},
                },
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )
        if is_nyc:
            # LLM contract: CAPTURED -> COMPLETE atomically advances
            # DISCOVERED -> ENCODED; ROUNDTRIP_VERIFIED -> BENCHMARKED -> REPORTED remain downstream.
            finalize_capture(destination, CaptureState.COMPLETE)
        succeeded = True
    except BaseException as error:
        if is_nyc:
            if (destination / "capture.json").is_file():
                fail_active_capture(destination, error)
        raise
    finally:
        if is_nyc and not succeeded:
            for path in (partial, output, destination / "materialization.json"):
                path.unlink(missing_ok=True)
    return output
